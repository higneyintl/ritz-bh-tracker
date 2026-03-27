#!/usr/bin/env python3
"""
Ritz BH Tracker — Sync Script
Reads the XLSX from Google Drive → generates data.js → pushes to GitHub → Netlify auto-deploys.
Run this after updating the spreadsheet.
"""

import subprocess
import sys
import os
import json
import shutil
import tempfile

DRIVE_FILE_ID   = "1BSGLcRtlHM0GCqgDD1tRZIizJ2a5WVmy"
GOG_PATH        = "/opt/homebrew/bin/gog"
REPO_DIR        = os.path.dirname(os.path.abspath(__file__))
DATA_JS         = os.path.join(REPO_DIR, "data.js")
XLSX_TMP        = os.path.join(tempfile.gettempdir(), "ritz_tracker_sync.xlsx")
GH_TOKEN        = open(os.path.expanduser("~/.openclaw/workspace/credentials/github-token.txt")).read().strip() if os.path.exists(os.path.expanduser("~/.openclaw/workspace/credentials/github-token.txt")) else os.environ.get("GH_TOKEN", "")

def log(msg, emoji="→"):
    print(f"  {emoji}  {msg}")

def run(cmd, cwd=None, capture=False):
    result = subprocess.run(cmd, shell=True, cwd=cwd,
                            capture_output=capture, text=True)
    if result.returncode != 0 and not capture:
        print(f"  ✗  Command failed: {cmd}")
        if capture:
            print(result.stderr)
    return result

# ── STEP 1: Download latest XLSX from Drive ───────────────────────────────
def download_xlsx():
    log("Downloading latest XLSX from Google Drive...", "⬇️")
    result = run(
        f'{GOG_PATH} drive download {DRIVE_FILE_ID} --output "{XLSX_TMP}" --account shameen@higneyintl.com',
        capture=True
    )
    if result.returncode != 0:
        # Try export as xlsx (if it was converted to Sheets)
        result2 = run(
            f'{GOG_PATH} drive export {DRIVE_FILE_ID} --mime "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" --output "{XLSX_TMP}" --account shameen@higneyintl.com',
            capture=True
        )
        if result2.returncode != 0:
            print("  ✗  Could not download file from Drive.")
            print(f"     {result2.stderr}")
            sys.exit(1)
    log(f"Downloaded to {XLSX_TMP}", "✅")

# ── STEP 2: Parse XLSX → data.js ─────────────────────────────────────────
def parse_xlsx():
    try:
        import openpyxl
    except ImportError:
        run("pip3 install openpyxl -q")
        import openpyxl

    log("Parsing XLSX...", "📊")
    wb = openpyxl.load_workbook(XLSX_TMP, data_only=True)

    if "Procurement Tracker" not in wb.sheetnames:
        print("  ✗  Sheet 'Procurement Tracker' not found.")
        print(f"     Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["Procurement Tracker"]
    items = []

    # Find header row (row 2)
    # Expected columns (1-indexed):
    # 1=VENDOR, 2=CODE, 3=DESCRIPTION, 4=QTY, 5=UNIT, 6=UNIT COST,
    # 7=EXT COST (formula), 8=SHIPPING (formula or "Incl."),
    # 9=TOTAL (formula), 10=HI COMM (formula),
    # 11=COMM PAID?, 12=PAID TO DATE, 13=BALANCE (formula),
    # 14=STATUS, 15=NOTES

    current_vendor = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        vendor = str(row[0]).strip() if row[0] else ""
        code   = str(row[1]).strip() if row[1] else ""
        desc   = str(row[2]).strip() if row[2] else ""

        # Skip vendor header rows (merged, code is blank)
        if vendor and not code:
            current_vendor = vendor
            continue

        # Skip empty rows
        if not vendor and not code:
            continue

        # Skip grand total row
        if vendor and "GRAND TOTAL" in vendor.upper():
            continue

        try:
            qty       = float(row[3]) if row[3] is not None else 0
            unit      = str(row[4]).strip() if row[4] else "ea"
            unit_cost = float(row[5]) if row[5] is not None else 0
        except (ValueError, TypeError):
            continue

        # Detect no_ship from column 8 value
        ship_val  = row[7]
        no_ship   = False
        if ship_val is None or str(ship_val).strip() in ("Incl.", "—", "N/A", ""):
            no_ship = True

        # Commission paid
        comm_paid_val = str(row[10]).strip() if row[10] is not None else "No"
        comm_paid = comm_paid_val.lower() in ("yes", "true", "1", "y")

        # Paid to date
        try:
            paid = float(row[11]) if row[11] is not None else 0
        except (ValueError, TypeError):
            paid = 0

        # Notes
        notes = str(row[14]).strip() if row[14] is not None else ""
        if notes in ("None", "none"):
            notes = ""

        items.append({
            "vendor":    vendor or current_vendor or "UNKNOWN",
            "code":      code,
            "desc":      desc,
            "qty":       qty,
            "unit":      unit,
            "unitCost":  unit_cost,
            "noShip":    no_ship,
            "paid":      paid,
            "commPaid":  comm_paid,
            "notes":     notes,
        })

    log(f"Parsed {len(items)} line items across {len(set(i['vendor'] for i in items))} vendors", "✅")
    return items

# ── STEP 3: Write data.js ─────────────────────────────────────────────────
def write_data_js(items):
    log("Writing data.js...", "✍️")

    lines = [
        "// ============================================================",
        "// RITZ CARLTON BAL HARBOUR — POOL #1004 — DATA FILE",
        "// AUTO-GENERATED by sync.py — do not edit directly",
        f"// Last synced: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "// To update: edit the XLSX in Google Drive, then run sync.py",
        "// ============================================================",
        "",
        "const BUDGET_DATA = [",
    ]

    current_vendor = None
    for item in items:
        if item["vendor"] != current_vendor:
            if current_vendor is not None:
                lines.append("")
            lines.append(f"  // {'─' * 54}")
            lines.append(f"  // {item['vendor']}")
            lines.append(f"  // {'─' * 54}")
            current_vendor = item["vendor"]

        no_ship_str = "true" if item["noShip"] else "false"
        comm_paid_str = "true" if item["commPaid"] else "false"
        desc_escaped = item["desc"].replace('"', '\\"').replace("'", "\\'")
        notes_escaped = item["notes"].replace('"', '\\"').replace("'", "\\'")

        lines.append(
            f'  {{vendor:"{item["vendor"]}", code:"{item["code"]}", desc:"{desc_escaped}", '
            f'qty:{item["qty"]}, unit:"{item["unit"]}", unitCost:{item["unitCost"]}, '
            f'noShip:{no_ship_str}, paid:{item["paid"]}, commPaid:{comm_paid_str}, '
            f'notes:"{notes_escaped}"}},'
        )

    lines += [
        "];",
        "",
        "const PROJECT_FIXED = {",
        "  storage_installation: 75000.00,",
        "  blondies_shipping_credit: -24528.00,",
        "  sales_tax_rate: 0.07,",
        "  shipping_rate: 0.10,",
        "  commission_rate: 0.10,",
        "};",
    ]

    with open(DATA_JS, "w") as f:
        f.write("\n".join(lines) + "\n")

    log(f"Written to {DATA_JS}", "✅")

# ── STEP 4: Git commit + push ─────────────────────────────────────────────
def git_push():
    log("Pushing to GitHub...", "🚀")

    remote_url = f"https://{GH_TOKEN}@github.com/higneyintl/ritz-bh-tracker.git"
    run(f'git remote set-url origin "{remote_url}"', cwd=REPO_DIR)
    run("git add data.js", cwd=REPO_DIR)

    from datetime import datetime
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    result = run(f'git diff --cached --quiet', cwd=REPO_DIR, capture=True)
    if result.returncode == 0:
        log("No changes detected — nothing to push.", "ℹ️")
        return False

    run(f'git commit -m "Sync from XLSX — {ts}"', cwd=REPO_DIR)
    run("git push origin main", cwd=REPO_DIR)
    log("Pushed! Netlify will deploy in ~30 seconds.", "✅")
    return True

# ── MAIN ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print()
    print("=" * 56)
    print("  RITZ BH TRACKER — SYNC")
    print("=" * 56)

    download_xlsx()
    items = parse_xlsx()
    write_data_js(items)
    pushed = git_push()

    print()
    print("=" * 56)
    if pushed:
        print("  ✅  DONE — Live at https://ritz-bh-tracker.netlify.app")
    else:
        print("  ℹ️   No changes — tracker is already up to date.")
    print("=" * 56)
    print()

    # Keep terminal open on Mac double-click
    if len(sys.argv) > 1 and sys.argv[1] == "--pause":
        input("  Press Enter to close...")
